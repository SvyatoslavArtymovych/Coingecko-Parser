from openpyxl import Workbook, load_workbook
from os import path


def create_table():
    if not path.exists('./Table.xlsx'):
        wb = Workbook()
        sheet = wb['Sheet']
        sheet.append(['Coin №', 'Name', 'Date', 'Link'])
        wb.save("./Table.xlsx")
        print('Table created')
    else:
        print('Table exists')


def write_empty_row():
    filename = './Table.xlsx'
    wb = load_workbook(filename=filename)
    sheet = wb['Sheet']
    sheet.append([''])
    wb.save(filename)


def write_information(info):
    filename = './Table.xlsx'
    wb = load_workbook(filename=filename)
    sheet = wb['Sheet']
    sheet.append(info)
    wb.save(filename)
